import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\redi_\Downloads\Uniformes.xlsx', data_only=True, read_only=True)

uniforms = []
# 1. READ INVENTARIO UNIFORMES
sheet_uni = wb['Inventario Uniformes ']
for row in sheet_uni.iter_rows(min_row=4, values_only=True):
    if not row[0] or str(row[0]).strip() == 'CONCEPTO': continue
    name = str(row[0]).strip()
    qty = float(row[1]) if row[1] is not None else 0
    cost = float(row[2]) if row[2] is not None else 0
    
    # Try to extract size from end of name (CH, M, G)
    size = None
    parts = name.split()
    if parts[-1] in ['CH', 'M', 'G', 'XG']:
        size = parts[-1]
        name = ' '.join(parts[:-1])
        
    uniforms.append({
        'name': name,
        'size': size,
        'category': 'uniforme',
        'qty': qty,
        'cost': cost
    })

# 2. READ INVENTARIO EPP
sheet_epp = wb['Inventario EPP']
for row in sheet_epp.iter_rows(min_row=4, values_only=True):
    if not row[0] or str(row[0]).strip() == 'CONCEPTO': continue
    name = str(row[0]).strip()
    qty = float(row[1]) if row[1] is not None else 0
    cost = float(row[2]) if row[2] is not None else 0
    
    uniforms.append({
        'name': name,
        'size': None,
        'category': 'epp',
        'qty': qty,
        'cost': cost
    })

sql = []
sql.append('-- 1. CREATE TABLES')
sql.append('''
CREATE TABLE inventory_uniforms (
    id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
    name VARCHAR(255) NOT NULL,
    category VARCHAR(20) NOT NULL,
    size VARCHAR(20),
    current_stock DECIMAL(10,2) DEFAULT 0,
    unit_cost DECIMAL(10,2) DEFAULT 0,
    is_active BOOLEAN DEFAULT true,
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE uniform_assignments (
    id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
    uniform_id UUID REFERENCES inventory_uniforms(id) ON DELETE CASCADE,
    employee_id UUID REFERENCES hr_employees(id) ON DELETE CASCADE,
    quantity DECIMAL(10,2) NOT NULL,
    assigned_date DATE NOT NULL,
    notes TEXT,
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TRIGGER update_inventory_uniforms_modtime BEFORE UPDATE ON inventory_uniforms FOR EACH ROW EXECUTE PROCEDURE update_modified_column();

ALTER TABLE inventory_uniforms ENABLE ROW LEVEL SECURITY;
ALTER TABLE uniform_assignments ENABLE ROW LEVEL SECURITY;
CREATE POLICY "Allow all access on inventory_uniforms" ON inventory_uniforms FOR ALL USING (true) WITH CHECK (true);
CREATE POLICY "Allow all access on uniform_assignments" ON uniform_assignments FOR ALL USING (true) WITH CHECK (true);
''')

sql.append('-- 2. INSERT UNIFORMES Y EPP')
for u in uniforms:
    size_val = f"'{u['size']}'" if u['size'] else 'NULL'
    sql.append(f"INSERT INTO inventory_uniforms (name, category, size, current_stock, unit_cost) VALUES ('{u['name']}', '{u['category']}', {size_val}, {u['qty']}, {u['cost']});")

# Write to file
with open('supabase/migrations/20260506010000_uniforms_epp.sql', 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql))

print('Migration file created successfully!')
