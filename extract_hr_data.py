import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(r'C:\Users\redi_\Downloads\Planificador de vacaiones .xlsm', data_only=True, read_only=True)

# 1. READ EMPLOYEES
sheet_col = wb['ListaCol']
employees = []
for row in sheet_col.iter_rows(min_row=6, values_only=True):
    if not row[1]: continue
    name = str(row[1]).strip()
    dept = str(row[2]).strip() if row[2] else ''
    hire_date = row[3] if isinstance(row[3], datetime) else None
    
    employees.append({
        'name': name,
        'dept': dept,
        'hire_date': hire_date.strftime('%Y-%m-%d') if hire_date else None
    })

# 2. READ ABSENCES
sheet_aus = wb['RegAus']
absences = []
for row in sheet_aus.iter_rows(min_row=2, values_only=True):
    if not row[1] or str(row[1]).strip() == 'NOMBRE': continue
    name = str(row[1]).strip()
    abs_type = str(row[2]).strip() if row[2] else 'VACACIONES'
    start_date = row[3] if isinstance(row[3], datetime) else None
    days = float(row[4]) if row[4] is not None else 0
    return_date = row[5] if isinstance(row[5], datetime) else None
    compensated = str(row[6]).strip().upper() == 'COMPENSADO' if row[6] else False
    
    if name and start_date:
        absences.append({
            'name': name,
            'type': abs_type,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'days': days,
            'return_date': return_date.strftime('%Y-%m-%d') if return_date else None,
            'is_compensated': compensated
        })

# Generate SQL
sql = []
sql.append('-- 1. CREATE TABLES')
sql.append('''
CREATE TABLE hr_employees (
    id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
    user_id UUID REFERENCES app_users(id) ON DELETE SET NULL,
    full_name VARCHAR(255) UNIQUE NOT NULL,
    department VARCHAR(100),
    hire_date DATE,
    base_vacation_days INTEGER DEFAULT 12,
    is_active BOOLEAN DEFAULT true,
    notes TEXT,
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE hr_absences (
    id UUID PRIMARY KEY DEFAULT uuid_generate_v4(),
    employee_id UUID REFERENCES hr_employees(id) ON DELETE CASCADE,
    absence_type VARCHAR(100) NOT NULL,
    start_date DATE NOT NULL,
    end_date DATE,
    days_count DECIMAL(5,1) NOT NULL,
    return_date DATE,
    is_compensated BOOLEAN DEFAULT false,
    notes TEXT,
    status VARCHAR(20) DEFAULT 'approved',
    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
);

CREATE TRIGGER update_hr_employees_modtime BEFORE UPDATE ON hr_employees FOR EACH ROW EXECUTE PROCEDURE update_modified_column();
CREATE TRIGGER update_hr_absences_modtime BEFORE UPDATE ON hr_absences FOR EACH ROW EXECUTE PROCEDURE update_modified_column();

ALTER TABLE hr_employees ENABLE ROW LEVEL SECURITY;
ALTER TABLE hr_absences ENABLE ROW LEVEL SECURITY;
CREATE POLICY "Allow all access on hr_employees" ON hr_employees FOR ALL USING (true) WITH CHECK (true);
CREATE POLICY "Allow all access on hr_absences" ON hr_absences FOR ALL USING (true) WITH CHECK (true);
''')

sql.append('-- 2. INSERT EMPLOYEES')
for emp in employees:
    hd = f"'{emp['hire_date']}'" if emp['hire_date'] else 'NULL'
    sql.append(f"INSERT INTO hr_employees (full_name, department, hire_date) VALUES ('{emp['name']}', '{emp['dept']}', {hd}) ON CONFLICT (full_name) DO NOTHING;")

sql.append('\n-- 3. INSERT ABSENCES')
for abs_rec in absences:
    rd = f"'{abs_rec['return_date']}'" if abs_rec['return_date'] else 'NULL'
    comp = 'true' if abs_rec['is_compensated'] else 'false'
    sql.append(f"INSERT INTO hr_absences (employee_id, absence_type, start_date, days_count, return_date, is_compensated) SELECT id, '{abs_rec['type']}', '{abs_rec['start_date']}', {abs_rec['days']}, {rd}, {comp} FROM hr_employees WHERE full_name = '{abs_rec['name']}';")

with open('supabase/migrations/20260506000000_hr_vacations.sql', 'w', encoding='utf-8') as f:
    f.write('\n'.join(sql))
print('Migration file created successfully!')
